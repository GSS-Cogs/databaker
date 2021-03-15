import sys
from datetime import datetime, time
import xlrd
from xlrd.biffh import XLRDError

from messytables.core import RowSet, TableSet, Cell, CoreProperties
from messytables.types import (StringType, IntegerType,
                               DateType, FloatType)
from messytables.error import ReadError
from messytables.compat23 import PY2

import openpyxl

class InvalidDateError(Exception):
    pass

def get_data_type(openpyexcel_cell):
    """
    Map the raw type of the openpyexcel cell to a messytable cell type
    """
    if isinstance(openpyexcel_cell, float):
        return FloatType()
    elif isinstance(openpyexcel_cell, time):
        return DateType(None)
    elif isinstance(openpyexcel_cell, int):
        return IntegerType()
    return StringType()

class XLSXTableSet(TableSet):
    """An excel workbook wrapper object.
    """

    def __init__(self, fileobj=None, filename=None, window=None,
                 encoding=None, with_formatting_info=True, **kw):

        def get_workbook():
            if fileobj is not None:
                return openpyxl.load_workbook(fileobj)
            return openpyxl.load_workbook(filename)

        self.window = window

        if not filename and not fileobj:
            raise Exception('You must provide one of filename or fileobj')

        self.workbook = get_workbook()
        if self.workbook.epoch == openpyxl.utils.datetime.CALENDAR_WINDOWS_1900:
            print("1900")
        else:
            print("Not 1900")

    def make_tables(self):
        """ Return the sheets in the workbook. """
        return [XLSXRowSet(name, self.workbook[name], self.window)
                for name in self.workbook.sheetnames]


class XLSXRowSet(RowSet):
    """ Excel support for a single sheet in the excel workbook. Unlike
    the CSV row set this is not a streaming operation. """

    def __init__(self, name, sheet, window=None):
        self.name = name
        self.sheet = sheet
        self.window = window or 1000
        super(XLSXRowSet, self).__init__(typed=True)

    def raw(self, sample=False):
        for row_no, row in enumerate(self.sheet.iter_rows()):
            row_of_cells = []
            for col_no, openpyexcel_cell in enumerate(row):
                row_of_cells.append(
                    XLSXCell.from_openpyexcel(openpyexcel_cell,
                            self.sheet, col_no, row_no))
            yield row_of_cells


class XLSXCell(Cell):

    @staticmethod
    def from_openpyexcel(openpyexcel_cell, sheet, col, row):
        value = openpyexcel_cell.value
        cell_type = get_data_type(openpyexcel_cell)
        
        from dateutil.parser import parse

        if cell_type == DateType(None):
            value = parse(value)

        messy_cell = XLSXCell(value, type=cell_type)
        messy_cell.sheet = sheet
        messy_cell.openpyexcel_cell = openpyexcel_cell
        messy_cell.xlrd_pos = (row, col) 

        return messy_cell

    @property
    def topleft(self):
        return self.properties.cell.topleft

    @property
    def properties(self):
        return XLSXProperties(self)

class XLSXProperties(CoreProperties):
    KEYS = ['bold', 'italic', 'underline', 'blank', 'strikeout', 
            'font_name', 'size', 'any_border', 'all_border', 
            'richtext', 'a_date', 'formatting_string']

    def __init__(self, cell):
        self.cell = cell
        self.merged = {}

    def get_bold(self):
        return self.cell.openpyexcel_cell.font.bold

    def get_italic(self):
        return self.cell.openpyexcel_cell.font.italic

    def get_underline(self):
        return self.cell.openpyexcel_cell.font.underline

    def get_blank(self):
        """Note that cells might not exist at all.
           Behaviour for spanned cells might be complicated: hence this function"""
        return self.cell.value == None
        #return self.cell.openpyexcel_cell.cell.value == ''
        
    def get_strikeout(self):
        return self.cell.openpyexcel_cell.font.strike

    def get_font_name(self):
        return self.cell.openpyexcel_cell.font.name

    def get_size(self):
        return self.cell.openpyexcel_cell.font.size

    def get_any_border(self):
        #b = self.cell.openpyexcel_cell.border
        #b = self.cell.openpyexcel_cell.style
        b = self.cell.properties
        print(b)

        #print(b.diagonalDown)
        #print(b.diagonalUp)
        #print(b.outline)

        #if b.style is not None:
        #    print(b.left or b.right or b.top or b.bottom )
        return b.diagonalDown or b.diagonalUp or b.outline
        #return b.left or b.right or b.top or b.bottom 
        #raise NotImplementedError('Any Border property not implemented for xlsx files.')

    def get_all_border(self):
        b = self.cell.openpyexcel_cell.border
        return b.left and b.right and b.top and b.bottom

    #def rich(self):
    #    return self.cell.
        #return self.cell.sheet.rich_text_runlist_map.get(self.cell.xlrd_pos, None)

    def get_richtext(self):
        return bool(self.rich)
